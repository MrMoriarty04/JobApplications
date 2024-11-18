from tkinter import *
import tkinter
import tkinter
from tkinter import ttk
from tkinter import messagebox
import os
import tkinter as tk  # Use "tk" as a shorter alias for cleaner code
import openpyxl
import openpyxl.workbook
import pandas as pd
import re
import datetime
from datetime import datetime
import time



def GUI(self):    
    def load_data():
        newpath = self 
        workbook = openpyxl.load_workbook(newpath)
        sheet = workbook.active

        list_values = list(sheet.values)
        print(list_values)
        for col_name in list_values[0]:
            treeview.heading(col_name, text=col_name)

        for value_tuple in list_values[1:]:
            treeview.insert('', tk.END, values=value_tuple)    
    root = tk.Tk()
    root.geometry("1400x900") 
    style = ttk.Style(root)
    root.tk.call("source", "forest-light.tcl")
    root.tk.call("source", "forest-dark.tcl")
    style.theme_use("forest-dark")
    frame = ttk.Frame(root)
    frame.pack()
    treeFrame = ttk.Frame(frame)
    treeFrame.grid(row=0, column=1, pady=10)
    treeScroll = ttk.Scrollbar(treeFrame)
    treeScroll.pack(side="right", fill="y")

    cols = ("First Name", "Mid Name", "Last Name", "Birth Date", "Gender","ID", "Major", "Level Of Education","GPA","Mobile Number", "Email","Accept Terms")
    treeview = ttk.Treeview(treeFrame, show="headings",
                            yscrollcommand=treeScroll.set, columns=cols, height=50)
    treeview.column("First Name", width=100)
    treeview.column("Mid Name", width=100)
    treeview.column("Last Name", width=100)
    treeview.column("Birth Date", width=100)
    treeview.column("Gender", width=100)
    treeview.column("ID", width=100)
    treeview.column("Major", width=100)
    treeview.column("Level Of Education", width=100)
    treeview.column("GPA", width=100)
    treeview.column("Mobile Number", width=100)
    treeview.column("Email", width=100)
    treeview.column("Accept Terms", width=100)

    treeview.pack()
    treeScroll.config(command=treeview.yview)
    load_data()
    root.mainloop()


def GUINew(self):    
    def load_data():
        newpath = self 
        workbook = openpyxl.load_workbook(newpath)
        sheet = workbook.active

        list_values = list(sheet.values)
        print(list_values)
        for col_name in list_values[0]:
            treeview.heading(col_name, text=col_name)

        for value_tuple in list_values[1:]:
            treeview.insert('', tk.END, values=value_tuple)    
    root = tk.Tk()
    root.geometry("1600x900") 
    style = ttk.Style(root)
    root.tk.call("source", "forest-light.tcl")
    root.tk.call("source", "forest-dark.tcl")
    style.theme_use("forest-dark")
    frame = ttk.Frame(root)
    frame.pack()
    treeFrame = ttk.Frame(frame)
    treeFrame.grid(row=0, column=1, pady=10)
    treeScroll = ttk.Scrollbar(treeFrame)
    treeScroll.pack(side="right", fill="y")

    cols = ("_Person__rank","_Person__fname", "_Person__mname", "_Person__lname","_Person__id", "_Person__birthDate" ,"_Person__gpa", "_Person__email", "_Person__gender","_Person__specialization","_Person__phone")
    treeview = ttk.Treeview(treeFrame, show="headings",
                            yscrollcommand=treeScroll.set, columns=cols, height=50)
    treeview.column("_Person__rank", width=130)
    treeview.column("_Person__fname", width=130)
    treeview.column("_Person__mname", width=130)
    treeview.column("_Person__lname", width=130)
    treeview.column("_Person__id", width=130)
    treeview.column("_Person__birthDate", width=130)
    treeview.column("_Person__gpa", width=130)
    treeview.column("_Person__email", width=130)
    treeview.column("_Person__gender", width=130)
    treeview.column("_Person__specialization", width=130)
    treeview.column("_Person__phone", width=130)
    # treeview.column("_Person__graduated", width=130)
    # treeview.column("Accept Terms", width=100)

    treeview.pack()
    treeScroll.config(command=treeview.yview)
    load_data()
    root.mainloop()    


def data_form():
    def enter_data():
        accepted =accept_var.get()
     

        # User info
        firstname = first_name_entry.get()
        first_name_entry.delete
        midname= mid_name_entry.get()
        lastname = last_name_entry.get()
        Birth= age_entry.get()
        Gender=gender_entry.get()
        ID_d=id_entry.get()
        if firstname and midname and lastname and Birth and Gender and ID_d:
            #edu
            Major=Major_entry.get()
            Education=degree_entry.get()
            GPA_s=GPA_entry.get()
            if Major and Education and GPA_s:
            #contact
                Mobile=phone_entry.get()
                Email=email_entry.get()
                if Mobile and Email:
                    filepath="/home/sherlock/Documents/Python/DATA.xlsx"
                    if not os.path.exists(filepath):
                        workbook=openpyxl.Workbook()
                        sheet= workbook.active
                        heading = ["First Name", "Mid Name", "Last Name", "Birth Date", "Gender","ID", "Major", "Level Of Education","GPA","Mobile Number", "Email","Accept Terms"]
                        sheet.append(heading)
                        workbook.save(filepath)
                    Workbook=openpyxl.load_workbook(filepath,read_only=False)
                    sheet=Workbook.active 
                    sheet.append([firstname, midname, lastname, Birth, Gender, ID_d, Major, Education, GPA_s, Mobile, Email,accepted])
                    Workbook.save(filepath)
                    print("Done!")
                    try:
                        workbook = openpyxl.load_workbook("/home/sherlock/Documents/Python/DATA.xlsx")
                        sheet = workbook.active  
                        print("Excel file loaded successfully!")
                        # Access and process data here (replace with your logic)
                        # For example, iterate through rows and columns:
                        for row in sheet.iter_rows():
                             for cell in row:
                                value = cell.value
                                print(value, end=" ")
                                print()
                    except FileNotFoundError:
                        print("Error: Excel file not found! Please check the filepath.")
                    except Exception as e:
                        print(f"An error occurred while reading the Excel file: {e}")
                    GUI(filepath)
                else:
                    tkinter.messagebox.showwarning(title= "Error", message="You have not enterd the contact correctly!")
            else:
                tkinter.messagebox.showwarning(title= "Error", message="You have not enterd the education correctly!")
        else:    
            tkinter.messagebox.showwarning(title= "Error", message="You have not enterd the user information correctly!")
    

    window = tk.Tk()
    window.title("Data Entry Form")
    window.geometry("1000x600") 


    style = ttk.Style(window)
    window.tk.call("source", "forest-light.tcl")
    window.tk.call("source", "forest-dark.tcl")
    style.theme_use("forest-dark")


    # Create a frame to hold the form elements
    frame = tk.Frame(window)
    frame.pack(padx=10, pady=10)  





     # Saving User Info
    user_info_frame =ttk.LabelFrame(frame, text="User Information")
    user_info_frame.grid(row= 0, column=0, padx=20, pady=10)

    first_name_label = ttk.Label(user_info_frame, text="First Name")
    first_name_label.grid(row=0, column=0)
    mid_name_label = ttk.Label(user_info_frame, text="Mid Name")
    mid_name_label.grid(row=0, column=1)
    last_name_label = ttk.Label(user_info_frame, text="Last Name")
    last_name_label.grid(row=0, column=2)

    first_name_entry = ttk.Entry(user_info_frame)
    mid_name_entry= ttk.Entry(user_info_frame)
    last_name_entry = ttk.Entry(user_info_frame)
    first_name_entry.grid(row=1, column=0)
    mid_name_entry.grid(row=1,column=1)
    last_name_entry.grid(row=1, column=2)



    age_label = ttk.Label(user_info_frame, text="Birth Date")
    age_entry = ttk.Entry(user_info_frame)
    age_label.grid(row=2, column=0)
    age_entry.grid(row=3, column=0)

    gender_label= ttk.Label(user_info_frame,text="Gender")
    gender_entry=ttk.Entry(user_info_frame)
    gender_label.grid(row=2,column=1)
    gender_entry.grid(row=3,column=1)

    id_label = ttk.Label(user_info_frame, text="ID")
    id_entry = ttk.Entry(user_info_frame)
    id_label.grid(row=2, column=2)
    id_entry.grid(row=3, column=2)

    for widget in user_info_frame.winfo_children():
         widget.grid_configure(padx=10, pady=5)




     # Saving Edu Info
    Edu_frame = ttk.LabelFrame(frame,text="Education")
    Edu_frame.grid(row=1, column=0, sticky="news", padx=20, pady=10)

    Major_label = ttk.Label(Edu_frame, text="Major")
    Major_entry = ttk.Entry(Edu_frame)
    Major_label.grid(row=0, column=0)
    Major_entry.grid(row=1, column=0)

    degree_label = ttk.Label(Edu_frame, text="Level Of Education")
    degree_entry = ttk.Entry(Edu_frame)
    degree_label.grid(row=0, column=1)
    degree_entry.grid(row=1, column=1)

    GPA_label = ttk.Label(Edu_frame, text="GPA")
    GPA_entry = ttk.Entry(Edu_frame)
    GPA_label.grid(row=0, column=2)
    GPA_entry.grid(row=1, column=2)

    for widget in Edu_frame.winfo_children():
        widget.grid_configure(padx=10, pady=5)

         # Saving contact Info
    courses_frame = ttk.LabelFrame(frame,text="contact")
    courses_frame.grid(row=2, column=0, sticky="news", padx=20, pady=10)

    phone_label = ttk.Label(courses_frame, text="Mobile Number")
    phone_entry = ttk.Entry(courses_frame)


    phone_label.grid(row=0, column=0)
    phone_entry.grid(row=1, column=0)


    email_label = ttk.Label(courses_frame, text="Email")
    email_entry = ttk.Entry(courses_frame)
    email_label.grid(row=0, column=1)
    email_entry.grid(row=1, column=1)




    for widget in courses_frame.winfo_children():
        widget.grid_configure(padx=10, pady=5)


    # Accept terms

    
    terms_frame = ttk.LabelFrame(frame, text="Terms & Conditions")
    terms_frame.grid(row=3, column=0, sticky="news", padx=20, pady=10)
    accept_var = tkinter.StringVar(value="Not Accepted")
    a = tk.BooleanVar()
    checkbutton = ttk.Checkbutton(frame, text="I accept the terms and conditions.", variable=a)
    checkbutton.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")
    # terms_check.grid(row=0, column=0)

    # Button
    button = ttk.Button(frame, text="Enter data", command= enter_data)
    button.grid(row=4, column=0, sticky="news", padx=20, pady=10)
    #Start the main event loop to display the window


    window.mainloop()


def data_show():

     
    def validateName(name):
            name = re.sub(r"(.)\1\1+", r"\1\1", name.lower())
            name = re.sub(r"[^A-Za-z]", "" ,name)
            return name[:25].capitalize() if re.search("^[A-Z][a-z]{1,25}$", name.capitalize()) else None

    def validateDate(date_str):
        try:
            date_format='%d-%m-%Y'
            date_struct = time.strptime(date_str, date_format)
            date_tuple = time.strptime(date_str, date_format)[:3]
            current_year = time.localtime().tm_year
            start_year = current_year - 60
            end_year = current_year - 20
            date_year = date_struct.tm_year

            if start_year <= date_year <= end_year:
                return True
            else:
                return False
        except ValueError:
            return False

    class Person:
        def __init__(self, fname, mname, lname, birthDate,gender,id,specialization, gpa,phone, email):
            self.__rank = 0
            self.setFname(fname)
            self.setMname(mname)
            self.setLname(lname)
            self.setId(id)
            self.setBirthDate(birthDate)
            self.setGpa(gpa)
            self.setEmail(email)
            self.setGender(gender)
            self.setSpecialization(specialization)
            self.setPhone(phone)
            self.setRank()

        # Rank
        def getRank(self):
            return self.__rank

        def setRank(self):
            if self.getFname() is not None or self.getMname() is not None or self.getLname() is not None:
                self.__rank += 5

            if self.getGpa() is not None and float(self.getGpa()) >= 3.5:
                self.__rank += 10

            specializationPattern = r"^(Cs|Cis|Se|Ai|Ds|It|Computer Science|Computer Information System|Software Engineer|Artificial Intelligence|Cyber Security|Data Science|Information Technology)$"
            if re.match(specializationPattern, self.getSpecialization()):
                self.__rank += 10

            if self.getEmail() is not None or self.getPhone() is not None:
                self.__rank += 5

            if self.getId() is None or (self.getEmail() is None and self.getPhone() is None):
                self.__rank = 0

        # First Name
        def getFname(self):
            return self.__fname

        def setFname(self, fname):
            self.__fname = None if validateName(fname) is None else validateName(fname)

        # Middle Name
        def getMname(self):
            return self.__mname

        def setMname(self, mname):
            self.__mname = None if validateName(mname) is None else validateName(mname)

        # Last Name
        def getLname(self):
            return self.__lname

        def setLname(self, lname):
            self.__lname = None if validateName(lname) is None else validateName(lname)

        # ID
        def getId(self):
            return self.__id

        def setId(self, id):
            self.__id = id if re.search(r"^(9(6[4-9]|[7-9][0-9])[01]|2000)\d{6}$", id) else None

        # Birth Date
        def getBirthDate(self):
            return self.__birthDate

        def setBirthDate(self, birthDate):
            self.__birthDate = birthDate


        # GPA
        def getGpa(self):
            return self.__gpa

        def setGpa(self, gpa):
            self.__gpa = (
                f"{float(gpa):.2f}" if re.search(r"^([1-3](\.[0-9]{1,4})?|4(\.00?)?)$", gpa)
                else f"{float(gpa)/25:.2f}" if re.search(r"^([5-9][0-9](\.[0-9]{1,2})?|100|0?\.[0-9]{1,2})$", gpa)
                else None
            )

        # Email
        def getEmail(self):
            return self.__email

        def setEmail(self, email):
            emailPattern = r"^([A-Za-z0-9]([-._]|[A-Za-z0-9])?){,31}[A-Za-z0-9]{1,2}@(([A-Za-z0-9](-[A-Za-z0-9])?)+\.([A-Za-z0-9](-[A-Za-z0-9])?)+)+$"
            self.__email = email if re.search(emailPattern, email) else None

        # Gender
        def getGender(self):
            return self.__gender

        def setGender(self, gender):
            self.__gender = (
                "Female" if re.search("^[FfEeGg][EeAa]*[MmNn]*[Aa]*[Ii]?(l|e|L|E|a|A)*$", gender)
                else "Male" if re.search("^[MmNn]+[Aa]*[Ii]?(l|e|L|E)*$", gender)
                else None
           )

        # Specialization
        def getSpecialization(self):
            return self.__specialization

        def setSpecialization(self, specialization):
            self.__specialization = ' '.join(word.capitalize() for word in specialization.split())

        # Phone
        def getPhone(self):
            return self.__phone

        def setPhone(self, phone):
            self.__phone = phone if re.search(r"^07[789][0-9]{7}$", phone) else None





    
    file="DATA.xlsx"
    D1filepath=file
    person_list = [] 

    df = pd.read_excel(D1filepath,header=None,skiprows=1)  # Skip 1 row (header)
    num_cols = len(df.columns)
    for index,row in df.iterrows():
        row_data=list(row)
        fname = row_data[0]
        mname = row_data[1]
        lname = row_data[2]
        birthDate = row_data[3]
        gender = row_data[4]
        id = row_data[5]
        specialization = row_data[6]
        dgree = row_data[7]
        gpa = row_data[8]
        phone = row_data[9]
        email = row_data[10]
        # graduated=row_data[11]
        person= Person(fname, mname, lname, birthDate,gender,id,specialization, gpa,phone, email)
        person_list.append(person)
        person_df = pd.DataFrame([p.__dict__ for p in person_list]) 
        new_file = "DATA2.xlsx"
        person_df.to_excel(new_file, index=False)
        print("Person objects stored in", new_file)
    GUINew(new_file)



window = tk.Tk()
window.title("Data Entry Form")
window.geometry("1000x600") 

style = ttk.Style(window)
window.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

frame = tk.Frame(window)
frame.pack(expand=True) 

button1 = Button(frame, text="Enter Data", padx=10, pady=5,command=data_form)
button1.pack(side=LEFT,padx=10)  # Place to the left within the frame
button2 = Button(frame, text="Show Data", padx=10, pady=5,command=data_show)
button2.pack(side=RIGHT,padx=10)  

window.mainloop()